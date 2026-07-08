"""Regression coverage for GET /api/steps/{id}/export (CSV + XLSX).

Invariants under test:
- CSV is emitted as UTF-8 **with a BOM** so Excel renders non-ASCII (e.g.
  Hebrew) headers/values instead of ANSI mojibake.
- XLSX is a real, openable workbook carrying the same Unicode data.
- A Unicode widget title never crashes the (latin-1-only) Content-Disposition
  header — it is delivered via the RFC 6266 filename* parameter.
- Unknown formats are rejected; unknown steps 404.

The step payload is seeded directly (steps are normally produced by the agent
loop); everything else goes through the real HTTP surface.
"""
import asyncio
import io
import uuid

import pytest
from openpyxl import load_workbook

from app.dependencies import async_session_maker
from app.models.widget import Widget
from app.models.query import Query
from app.models.step import Step

# Hebrew headers + values — the exact non-ASCII case that exposed the bug.
HEB_HEADER = "מכר נטו"          # "net sales"
HEB_VALUE = "תל אביב"          # "Tel Aviv"
COLUMNS = [
    {"field": "store", "headerName": "חנות"},
    {"field": "net", "headerName": HEB_HEADER},
    {"field": "qty", "headerName": "כמות"},
]
ROWS = [
    {"store": HEB_VALUE, "net": 15794415, "qty": 1360058},
    {"store": "ירושלים", "net": 15422305, "qty": 1319036},
]


def _auth(token, org_id):
    return {"Authorization": f"Bearer {token}", "X-Organization-Id": str(org_id)}


async def _seed_step(report_id, org_id, user_id, title="סך המכירות"):
    """Insert a widget + query + step (with Hebrew data) under a report."""
    suffix = uuid.uuid4().hex[:8]
    async with async_session_maker() as db:
        widget = Widget(title=title, slug=f"w-{suffix}", report_id=report_id)
        db.add(widget)
        await db.flush()

        query = Query(
            title=title, report_id=report_id, widget_id=widget.id,
            organization_id=org_id, user_id=user_id,
        )
        db.add(query)
        await db.flush()

        step = Step(
            title=title, slug=f"s-{suffix}", status="success",
            widget_id=widget.id, query_id=query.id,
            code="def generate_df(...):\n    return df",
            data={"rows": ROWS, "columns": COLUMNS},
            data_model={"type": "table", "columns": COLUMNS},
        )
        db.add(step)
        await db.flush()
        widget.default_step_id = step.id
        await db.commit()
        return step.id


@pytest.fixture
def seeded_step(create_user, login_user, whoami, create_report):
    user = create_user()
    token = login_user(user["email"], user["password"])
    me = whoami(token)
    org_id = me["organizations"][0]["id"]
    report = create_report(title="Export Report", user_token=token, org_id=org_id)
    step_id = asyncio.run(_seed_step(report["id"], org_id, me["id"]))
    return {"token": token, "org_id": org_id, "step_id": step_id}


@pytest.mark.e2e
class TestStepExport:
    def test_csv_has_bom_and_unicode(self, test_client, seeded_step):
        resp = test_client.get(
            f"/api/steps/{seeded_step['step_id']}/export",
            headers=_auth(seeded_step["token"], seeded_step["org_id"]),
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/csv")

        body = resp.content
        # BOM is the whole point — without it Excel mis-decodes Hebrew.
        assert body[:3] == b"\xef\xbb\xbf"
        text = body.decode("utf-8-sig")
        assert HEB_HEADER in text
        assert HEB_VALUE in text

    def test_csv_is_default_format(self, test_client, seeded_step):
        explicit = test_client.get(
            f"/api/steps/{seeded_step['step_id']}/export?format=csv",
            headers=_auth(seeded_step["token"], seeded_step["org_id"]),
        )
        assert explicit.status_code == 200
        assert explicit.headers["content-type"].startswith("text/csv")

    def test_xlsx_is_valid_workbook_with_unicode(self, test_client, seeded_step):
        resp = test_client.get(
            f"/api/steps/{seeded_step['step_id']}/export?format=xlsx",
            headers=_auth(seeded_step["token"], seeded_step["org_id"]),
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header_row = [c.value for c in ws[1]]
        assert HEB_HEADER in header_row
        # Numbers stay numeric in xlsx (not stringified).
        all_values = [c.value for row in ws.iter_rows() for c in row]
        assert HEB_VALUE in all_values
        assert 15794415 in all_values

    def test_unicode_title_does_not_break_content_disposition(self, test_client, seeded_step):
        # A Hebrew widget title must not crash the latin-1 header; the real name
        # rides in filename* (RFC 6266) with an ASCII fallback in filename.
        for fmt in ("csv", "xlsx"):
            resp = test_client.get(
                f"/api/steps/{seeded_step['step_id']}/export?format={fmt}",
                headers=_auth(seeded_step["token"], seeded_step["org_id"]),
            )
            assert resp.status_code == 200
            cd = resp.headers["content-disposition"]
            assert "filename*=UTF-8''" in cd
            assert cd.endswith(f".{fmt}")

    def test_unsupported_format_rejected(self, test_client, seeded_step):
        resp = test_client.get(
            f"/api/steps/{seeded_step['step_id']}/export?format=pdf",
            headers=_auth(seeded_step["token"], seeded_step["org_id"]),
        )
        assert resp.status_code == 400

    def test_unknown_step_404(self, test_client, seeded_step):
        resp = test_client.get(
            f"/api/steps/{uuid.uuid4()}/export",
            headers=_auth(seeded_step["token"], seeded_step["org_id"]),
        )
        assert resp.status_code == 404
